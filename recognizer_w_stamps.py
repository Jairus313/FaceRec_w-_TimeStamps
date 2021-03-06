from PIL import Image
import cv2
import tensorflow as tf
from keras.models import load_model
import numpy as np

import datetime
import openpyxl
import time

model = tf.keras.models.load_model('model.h5')
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

wb = openpyxl.Workbook() 
sheet = wb.active 

cell_1 = sheet['A1'] 
cell_1.value = "Name"
  
cell_2 = sheet['B1'] 
cell_2.value = "Time"

cell = 2

def face_extractor(img):
    faces = face_cascade.detectMultiScale(img, 1.3, 5)
    
    if faces is ():
        return None
    
    # Crop all faces found
    for (x,y,w,h) in faces:
        cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,255),2)
        cropped_face = img[y:y+h, x:x+w]

    return cropped_face

video_capture = cv2.VideoCapture("test.mp4")
while True:
    _, frame = video_capture.read()
  
    face=face_extractor(frame)
    if type(face) is np.ndarray:
        face = cv2.resize(face, (224, 224))
        im = Image.fromarray(face, 'RGB')
        img_array = np.array(im)
        img_array = np.expand_dims(img_array, axis=0)
        test = np.float32(img_array)
        pred = model.predict(test)
        print(pred)
                     
        name="None matching"
        
        if(pred[0][0]>0.5):
            name='Steve Roger'
            cv2.putText(frame,name, (50, 50), cv2.FONT_HERSHEY_COMPLEX, 1, (0,255,0), 2)

            datetime_object = datetime.datetime.now()
            val_x = sheet.cell(row = cell, column = 1) 
            val_x.value = name
            val_y = sheet.cell(row = cell , column = 2) 
            val_y.value = datetime_object
            cell += 1

        if(pred[0][1]>0.5):
            name='Sudeep Nellur'
            cv2.putText(frame,name, (50, 50), cv2.FONT_HERSHEY_COMPLEX, 1, (0,255,0), 2)

            datetime_object = datetime.datetime.now()
            val_x = sheet.cell(row = cell, column = 1) 
            val_x.value = name
            val_y = sheet.cell(row = cell , column = 2) 
            val_y.value = datetime_object
            cell += 1

    else:
        cv2.putText(frame,"No face found", (50, 50), cv2.FONT_HERSHEY_COMPLEX, 1, (0,255,0), 2)
    cv2.imshow('Video', frame)

    wb.save("timestamps.csv") 

    if cv2.waitKey(25)  == 13:
        break
    
video_capture.release()
cv2.destroyAllWindows()